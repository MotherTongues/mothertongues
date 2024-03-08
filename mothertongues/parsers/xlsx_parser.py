from typing import Any, Dict, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils.exceptions import InvalidFileException

from mothertongues.config.models import DataSource
from mothertongues.exceptions import ParserError, UnsupportedFiletypeError
from mothertongues.parsers import BaseTabularParser
from mothertongues.parsers.utils import col2int


class Parser(BaseTabularParser):
    def __init__(self, data_source: DataSource):
        self.resource_path = data_source.resource
        self.manifest = data_source.manifest
        try:
            work_book = load_workbook(self.resource_path, data_only=True)
        except InvalidFileException:
            raise UnsupportedFiletypeError(self.resource_path)
        if self.manifest.sheet_name is not None:
            try:
                work_sheet = work_book[self.manifest.sheet_name]
            except KeyError as e:
                raise ParserError(
                    f"{self.manifest.sheet_name} does not exist in the workbook at {self.resource_path}"
                ) from e
        else:
            work_sheet = work_book.active
        self.fieldnames: Optional[Dict[str, Any]] = None
        if self.manifest.use_header:
            self.fieldnames = {
                c.value: c.column
                for (c,) in work_sheet.iter_cols(min_row=1, max_row=1)
            }
            min_row = 2
        elif self.manifest.skip_header:
            min_row = 2
        else:
            min_row = 1
        self.resource = work_sheet.iter_rows(min_row=min_row)

    def parse_fn(self, entry: Tuple[Cell, ...], col: str) -> str:
        """Given a tuple of OpenPyxl cells, return the value of the cell matching the column value for col"""
        if self.fieldnames is not None and col in self.fieldnames:
            columns = [self.fieldnames[col], col2int(col)]
        else:
            columns = [col, col2int(col)]
        for c in entry:
            if c.column in columns:
                if isinstance(c.value, float):
                    return str(int(c.value)) if c.value.is_integer() else str(c.value)
                else:
                    return c.value if c.value is not None else ""
        return ""
