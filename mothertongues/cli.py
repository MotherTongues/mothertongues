import json
import shutil
import socketserver
from functools import partial
from http.server import SimpleHTTPRequestHandler
from pathlib import Path

import nltk
import questionary
import typer
from loguru import logger
from openpyxl import Workbook
from rich import print
from rich.padding import Padding
from rich.panel import Panel

from mothertongues.config import SchemaTypes, get_schemas
from mothertongues.config.models import (
    SCHEMA_DIR,
    DataSource,
    LanguageConfiguration,
    MTDConfiguration,
    ParserEnum,
    ParserTargets,
    ResourceManifest,
)
from mothertongues.dictionary import MTDictionary
from mothertongues.utils import load_mtd_configuration

app = typer.Typer(rich_markup_mode="markdown")
UI_DIR = Path(__file__).parent / "ui"


@app.command()
def run(
    dictionary_data: Path = typer.Argument(
        exists=True,
        file_okay=True,
        dir_okay=False,
        readable=True,
        help="The path to your generated dictionary data",
    ),  # type: ignore
    port: int = 3636,
):
    """
    ## Run your dictionary in your browser\

    This is helpful for debugging, just run this command and your dictionary will be available at [http://localhost:3636](http://localhots:3636)
    You can exit the server by pressing ctrl+c.

    If you haven't generated your dictionary data yet, use the export command first or use the build-and-run command instead.
    """
    shutil.copy(dictionary_data, UI_DIR / "assets" / "dictionary_data.json")
    Handler = partial(SimpleHTTPRequestHandler, directory=UI_DIR)
    logger.warning(
        "WARNING: This is a Development server and is not secure for production"
    )
    httpd = socketserver.TCPServer(("", port), Handler)
    logger.info(
        "Open http://localhost:{port} in your browser to see your dictionary", port=port
    )
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        httpd.server_close()


@app.command()
def build_and_run(
    language_config_path: Path = typer.Argument(
        exists=True,
        file_okay=True,
        dir_okay=False,
        readable=True,
        help="The path to your dictionary's language configuration file.",
    ),  # type: ignore
    port: int = 3636,
):
    """
    ## Build your dictionary and run it in your browser\

    This is helpful for debugging, just run this command and your dictionary will be available at [http://localhost:3636](http://localhots:3636)
    You can exit the server by pressing ctrl+c.
    """
    config = MTDConfiguration(**load_mtd_configuration(language_config_path))
    dictionary = MTDictionary(config)
    output = dictionary.export()
    Handler = partial(SimpleHTTPRequestHandler, directory=UI_DIR)
    with open(UI_DIR / "assets" / "dictionary_data.json", "w", encoding="utf8") as f:
        json.dump(output.model_dump(mode="json"), f, indent=4)
    logger.warning(
        "WARNING: This is a Development server and is not secure for production"
    )
    httpd = socketserver.TCPServer(("", port), Handler)
    logger.info(
        "Open http://localhost:{port} in your browser to see your dictionary", port=port
    )
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        httpd.server_close()


@app.command(hidden=True)
def schema(
    type: SchemaTypes = typer.Argument(
        default=SchemaTypes.main_format,
        help="The SchemaType to return, choose from the main format, or just the dictionary configuration, or the schema for each entry in the dictionary.",
    ),
    output: Path = typer.Argument(
        exists=False,
        file_okay=True,
        dir_okay=False,
        help="The file path to write the JSON schema",
    ),  # type: ignore
):
    """
    ## Export the JSON Schema for various MotherTongues Dictionary data definitions\

    This is helpful when creating your own UI for use with MTD-generated data. Read more about it here: [https://docs.mothertongues.org/](https://docs.mothertongues.org/)

    """
    schema = get_schemas(type)
    with open(output, "w", encoding="utf8") as f:
        json.dump(schema, f, indent=4)
    return schema


@app.command(hidden=True)
def update_schemas():
    """
    ## Update the packaged version of the schemas, this is useful in development.

    Unless you are an advanced user of mothertongues, do not worry about this command.

    """
    manifest_schema = json.loads(get_schemas(SchemaTypes.manifest, json=True))
    # manifest_schema['properties']['$schema'] = {
    #         "title": "Schema Path",
    #         "type": "string"
    #     }
    config_schema = json.loads(get_schemas(SchemaTypes.config, json=True))
    # config_schema['properties']['$schema'] = {
    #         "title": "Schema Path",
    #         "type": "string"
    #     }
    with open(SCHEMA_DIR / "manifest.json", "w", encoding="utf8") as f:
        json.dump(manifest_schema, f, indent=4)
    with open(SCHEMA_DIR / "config.json", "w", encoding="utf8") as f:
        json.dump(config_schema, f, indent=4)


@app.command()
def export(
    language_config_path: Path = typer.Argument(
        exists=True,
        file_okay=True,
        dir_okay=False,
        readable=True,
        help="The path to your dictionary's language configuration file.",
    ),  # type: ignore
    output_directory: Path = typer.Argument(
        exists=True,
        file_okay=False,
        dir_okay=True,
        help="The output directory to write to.",
    ),  # type: ignore
    include_info: bool = True,
):
    """
    ## Export your dictionary for use in a MTD UI\

    You need to export your data to create the file necessary for the UI. Read more about it here: [https://docs.mothertongues.org/](https://docs.mothertongues.org/)
    """
    config = MTDConfiguration(**load_mtd_configuration(language_config_path))
    dictionary = MTDictionary(config)
    if not dictionary.data:
        logger.error(
            "The 'data' value in your json config is empty. Cannot proceed with export."
        )
        print(
            Padding(
                Panel(
                    "",
                    title="[red]Nothing here, your dictionary is empty!",
                    subtitle="Please check your data and configurations.",
                ),
                (2, 4),
            )
        )
        exit(1)
    output = dictionary.export()
    if include_info:
        dictionary.print_info()
    logger.info(
        f"Writing dictionary data file to {(output_directory / 'dictionary_data.json')}"
    )
    with open(output_directory / "dictionary_data.json", "w", encoding="utf8") as f:
        json.dump(output.model_dump(mode="json"), f, indent=4)


@app.command()
def new_project(
    outdir: Path = typer.Option(
        default=None,
        file_okay=False,
        dir_okay=True,
        help="The output directory to write to.",
    ),
    overwrite: bool = typer.Option(default=False, help="Overwrite any existing files."),
):
    """Create a start project with default configurations and some mock data

    Args:
        outdir (Path, optional): The output directory to write to. Defaults to typer.Option( default=None, file_okay=False, dir_okay=True, help="The output directory to write to.", ).
        overwrite (bool, optional): _description_. Defaults to typer.Option(default=False).
    """
    if outdir is None:
        outdir = Path(
            questionary.path(
                "Where would you like to save your project?", only_directories=True
            ).ask()
        )
    outdir.mkdir(exist_ok=True)
    # Create the sample resource manifest and configuration
    resource_manifest = ResourceManifest(
        file_type=ParserEnum.xlsx, targets=ParserTargets(word="A", definition="B")
    )
    resource_path = outdir / "data.xlsx"
    config_path = outdir / "config.mtd.json"
    if resource_path.exists() and not overwrite:
        logger.error(
            f"Tried to generate sample data at {resource_path} but it already exists. Please choose another location or re-run with the --overwrite flag"
        )
        exit(1)
    if config_path.exists() and not overwrite:
        logger.error(
            f"Tried to generate configuration file at {config_path} but it already exists. Please choose another location or re-run with the --overwrite flag"
        )
        exit(1)
    # Create and write the sample data
    wb = Workbook()
    ws = wb.active
    # TODO: get more reasonable example data
    nltk.download("brown")
    for i, sent in enumerate(nltk.corpus.brown.words()[:50]):
        ws[f"A{i+1}"] = sent
        ws[f"B{i+1}"] = sent
    wb.save(resource_path)
    config = MTDConfiguration(
        config=LanguageConfiguration(sorting_field="word"),
        data=[DataSource(manifest=resource_manifest, resource=resource_path)],
    )
    # TODO: This adds a path that gets incorrectly resolved otherwise
    config.data[0].resource = "data.xlsx"  # type: ignore
    # Ignore the type check because model_dump is provided by the inherited pydantic class
    config_json = config.model_dump(exclude_none=True, mode="json")  # type: ignore
    # write the configuration files
    with open(config_path, "w", encoding="utf8") as f:
        json.dump(config_json, f, indent=4)

    print("")
    print("COMPLETE: MotherTongues project created with sample data")
    print(
        "You can find your configuration and data files here: " + str(outdir.absolute())
    )


if __name__ == "__main__":
    app()
